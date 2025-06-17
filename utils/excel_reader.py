import openpyxl
from datetime import datetime
def read_test_cases(filepath):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        testcase, username, password, module, action, expected, extra_data = row
        data.append({
            "testcase": testcase,
            "username": username,
            "password": password,
            "module": module,
            "action": action,
            "expected": expected,
            "extra_data": extra_data
        })

    return data

def format_date(value):
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    return str(value)  # assume it's already a string

def get_leave_data(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append({
            "username": row[0],
            "password": row[1],
            "leave_type": row[2],
            "from_date": format_date(row[3]),
            "to_date": format_date(row[4]),
            "comment": row[5]
        })
    return data